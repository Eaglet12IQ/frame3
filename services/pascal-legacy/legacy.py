#!/usr/bin/env python3
import os
import random
import time
import datetime
import csv
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

def get_env(key, default):
    return os.getenv(key, default)

def generate_and_insert():
    # Output directory
    out_dir = get_env('CSV_OUT_DIR', '/data/csv')
    os.makedirs(out_dir, exist_ok=True)

    # Generate filename
    now = datetime.datetime.now()
    timestamp_str = now.strftime('%Y%m%d_%H%M%S')
    csv_filename = f'telemetry_{timestamp_str}.csv'
    xlsx_filename = f'telemetry_{timestamp_str}.xlsx'
    csv_path = os.path.join(out_dir, csv_filename)
    xlsx_path = os.path.join(out_dir, xlsx_filename)

    # Generate data
    recorded_at_unix = int(time.time())
    voltage = round(3.2 + random.random() * (12.6 - 3.2), 2)
    temp = round(-50 + random.random() * 130, 2)
    is_valid = random.choice([True, False])
    source_file = csv_filename

    # Generate CSV
    with open(csv_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['recorded_at', 'voltage', 'temp', 'is_valid', 'source_file'])
        writer.writerow([recorded_at_unix, voltage, temp, 'TRUE' if is_valid else 'FALSE', source_file])

    # Generate XLSX with formulas
    wb = Workbook()
    ws = wb.active
    ws.title = 'Telemetry'

    # Headers
    ws['A1'] = 'recorded_at'
    ws['B1'] = 'voltage'
    ws['C1'] = 'temp'
    ws['D1'] = 'is_valid'
    ws['E1'] = 'source_file'

    # Data with formulas
    ws['A2'] = '=NOW()'  # Value substitution for date/time
    ws['B2'] = f'=3.2 + RAND() * {12.6 - 3.2}'  # Voltage with RAND
    ws['C2'] = f'=-50 + RAND() * 130'  # Temp with RAND
    ws['D2'] = '=TRUE' if is_valid else '=FALSE'  # Logical
    ws['E2'] = f'="telemetry_" & TEXT(NOW(), "yyyymmdd_hhmmss") & ".xlsx"'  # Dynamic filename

    # Style for datetime
    date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD HH:MM:SS')
    wb.add_named_style(date_style)
    ws['A2'].style = date_style

    wb.save(xlsx_path)

    # Database insertion
    host = get_env('PGHOST', 'db')
    port = get_env('PGPORT', '5432')
    user = get_env('PGUSER', 'monouser')
    password = get_env('PGPASSWORD', 'monopass')
    dbname = get_env('PGDATABASE', 'monolith')

    conn = psycopg2.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        dbname=dbname
    )
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO telemetry_legacy (recorded_at, voltage, temp, source_file, is_valid) VALUES (%s, %s, %s, %s, %s)",
        (now, voltage, temp, source_file, is_valid)
    )
    conn.commit()
    cur.close()
    conn.close()

    print(f'[python] Generated {csv_filename} and {xlsx_filename}, inserted into DB')

def main():
    random.seed()
    period_sec = int(get_env('GEN_PERIOD_SEC', '300'))

    print(f'[python] Legacy generator started (period = {period_sec} sec)')

    while True:
        generate_and_insert()
        print(f'[python] {datetime.datetime.now().strftime("%H:%M:%S")} — done')
        time.sleep(period_sec)

if __name__ == '__main__':
    main()
